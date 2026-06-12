#!/usr/bin/env python3
"""Build '05 - AAS Financial Projections.xlsx' — 36-month formula-driven model."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "05 - AAS Financial Projections.xlsx")

INK, PAPER, BLUE_BG, RULE = "0E1116", "F9F7F1", "DCE6FF", "E4E1D5"
F_INPUT  = Font(name="Arial", size=10, color="0000FF")           # blue = editable input
F_CALC   = Font(name="Arial", size=10, color="000000")           # black = formula
F_LINK   = Font(name="Arial", size=10, color="008000")           # green = cross-sheet link
F_LABEL  = Font(name="Arial", size=10, color="000000")
F_BOLD   = Font(name="Arial", size=10, bold=True, color="000000")
F_HDR    = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_TITLE  = Font(name="Arial", size=14, bold=True, color="000000")
F_SUB    = Font(name="Arial", size=9, italic=True, color="666666")
FILL_HDR = PatternFill("solid", start_color=INK)
FILL_IN  = PatternFill("solid", start_color=BLUE_BG)
FILL_SOFT= PatternFill("solid", start_color=PAPER)
THIN     = Border(bottom=Side(style="thin", color=RULE))
CUR      = '$#,##0;($#,##0);"-"'
CUR2     = '$#,##0.00;($#,##0.00);"-"'
PCT      = '0.0%;(0.0%);"-"'
NUM      = '#,##0;(#,##0);"-"'

MONTHS = 36
C0 = 3  # first month column = C
MLAB = []
mm, yy = 7, 2026  # model starts July 2026
NAMES = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
for i in range(MONTHS):
    MLAB.append(f"{NAMES[mm-1]}-{str(yy)[2:]}")
    mm += 1
    if mm == 13: mm, yy = 1, yy + 1

def col(i):  # month index (1-based) -> column letter
    return get_column_letter(C0 + i - 1)

LASTC = col(MONTHS)            # AL
Y1 = f"C:{col(12)}"            # C..N
YCOLS = [(3, 14), (15, 26), (27, 38)]  # numeric col ranges per year

def month_header(ws, row=2, title_row=3):
    ws.cell(row=row, column=2, value="Month #").font = F_SUB
    for i in range(1, MONTHS + 1):
        c = ws.cell(row=row, column=C0 + i - 1, value=i); c.font = F_SUB
        c.alignment = Alignment(horizontal="center")
        c2 = ws.cell(row=title_row, column=C0 + i - 1, value=MLAB[i - 1])
        c2.font = F_BOLD; c2.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(C0 + i - 1)].width = 9.5
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.freeze_panes = "C4"

def label(ws, r, text, bold=False, indent=False):
    c = ws.cell(row=r, column=2, value=("    " + text) if indent else text)
    c.font = F_BOLD if bold else F_LABEL
    return c

def put_row(ws, r, vals_or_formula, font, fmt=CUR, fill=None):
    for i in range(1, MONTHS + 1):
        v = vals_or_formula(i) if callable(vals_or_formula) else vals_or_formula[i - 1]
        c = ws.cell(row=r, column=C0 + i - 1, value=v)
        c.font = font; c.number_format = fmt
        if fill: c.fill = fill

def year_sums(ws, r, src_row=None, fmt=CUR, font=F_CALC):
    src_row = src_row or r
    for k, (a, b) in enumerate(YCOLS):
        cc = get_column_letter(40 + k)
        c = ws.cell(row=r, column=40 + k,
                    value=f"=SUM({get_column_letter(a)}{src_row}:{get_column_letter(b)}{src_row})")
        c.font = font; c.number_format = fmt
def year_headers(ws, row=3):
    for k, t in enumerate(["Year 1", "Year 2", "Year 3"]):
        c = ws.cell(row=row, column=40 + k, value=t)
        c.font = F_BOLD; c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(40 + k)].width = 12

wb = Workbook()

# ---------------------------------------------------------------- README
ws = wb.active; ws.title = "README"
ws.sheet_properties.tabColor = INK
ws.column_dimensions["B"].width = 100
rows = [
 ("AAS FINANCIAL PROJECTIONS — 36-MONTH OPERATING MODEL", F_TITLE),
 ("Advanced Agentic Systems · Prepared June 2026 · Model start: July 2026", F_SUB),
 ("", None),
 ("HOW TO USE", F_BOLD),
 ("1.  Edit BLUE cells only (Assumptions, Volumes, Startup Costs). Everything else recalculates.", F_LABEL),
 ("2.  Volumes tab holds the unit plan: engagements per month by offer. It is pre-filled with the base case.", F_LABEL),
 ("3.  P&L, Cash Flow, Break-Even and Dashboard are fully formula-driven — do not type over black or green cells.", F_LABEL),
 ("4.  Scenario view: Break-Even & Scenarios tab applies the conservative / stretch multipliers from Assumptions.", F_LABEL),
 ("", None),
 ("COLOR KEY", F_BOLD),
 ("BLUE text on light-blue fill = your editable inputs and planning assumptions.", F_INPUT),
 ("Black text = calculated on this sheet.  Green text = pulled from another sheet.", F_LABEL),
 ("", None),
 ("NOTES", F_BOLD),
 ("All figures are founder planning assumptions (sample placeholders), not guarantees of performance.", F_LABEL),
 ("Owner draw is modeled as an expense for conservatism. The LLC is a pass-through entity; confirm tax treatment with a CPA.", F_LABEL),
 ("Tax set-aside (Cash Flow) reserves 25% of positive monthly net income — adjust on the Assumptions tab with your CPA.", F_LABEL),
 ("Pilot revenue is recognized 60% in the start month and 40% the following month (editable on Assumptions).", F_LABEL),
]
for i, (t, f) in enumerate(rows, start=2):
    c = ws.cell(row=i, column=2, value=t)
    if f: c.font = f

# ---------------------------------------------------------------- Assumptions
ws = wb.create_sheet("Assumptions")
ws.sheet_properties.tabColor = "1E5BFF"
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 40
for cl in "CDE": ws.column_dimensions[cl].width = 14
ws.column_dimensions["F"].width = 56
ws.cell(row=2, column=2, value="ASSUMPTIONS — edit blue cells").font = F_TITLE

def arow(r, name, val, fmt=CUR, note="", formula=False):
    label(ws, r, name)
    c = ws.cell(row=r, column=3, value=val)
    c.font = F_CALC if formula else F_INPUT
    if not formula: c.fill = FILL_IN
    c.number_format = fmt
    if note:
        n = ws.cell(row=r, column=6, value=note); n.font = F_SUB
    return c

ws.cell(row=4, column=2, value="Pricing (per engagement)").font = F_BOLD
arow(5,  "Workshop (executive / operator session)", 2000, CUR, "Source: BUSINESS_GAME_PLAN.md — low-friction paid entry")
arow(6,  "Foundation — Workflow Intelligence Audit", 25000, CUR, "Source: BUSINESS_GAME_PLAN.md — fixed-fee anchor")
arow(7,  "Momentum — Agentic Workflow Pilot", 55000, CUR, "Midpoint of $40K–$90K typical range; success component not modeled (conservative)")
arow(8,  "Pilot revenue % in start month", 0.6, PCT)
label(ws, 9, "Pilot revenue % in following month")
c = ws.cell(row=9, column=3, value="=1-C8"); c.font = F_CALC; c.number_format = PCT
arow(10, "Operating Partner retainer (per month)", 6000, CUR, "Midpoint of $4K–$12K/mo posture")
arow(11, "Agency growth-system build", 15000, CUR, "Brand/web/chat/voice/CRM build, commercial")
arow(12, "Agency care plan (per month)", 1500, CUR)
arow(13, "Training & onboarding program", 12000, CUR)

ws.cell(row=15, column=2, value="Delivery cost (COGS, % of revenue)").font = F_BOLD
arow(16, "Year 1", 0.12, PCT, "Subcontract delivery support")
arow(17, "Year 2", 0.15, PCT)
arow(18, "Year 3", 0.15, PCT)

ws.cell(row=20, column=2, value="Owner draw").font = F_BOLD
arow(21, "Draw per month — Year 1", 4000, CUR)
arow(22, "Draw starts in month #", 4, NUM)
arow(23, "Draw per month — Year 2", 7000, CUR)
arow(24, "Draw per month — Year 3", 9000, CUR)

ws.cell(row=26, column=2, value="Hires (loaded monthly cost incl. taxes/benefits)").font = F_BOLD
hires = [("Delivery analyst", 14, 7800), ("Automation engineer", 20, 11450),
         ("Client success lead", 26, 9375), ("Product engineer", 32, 12500)]
ws.cell(row=27, column=3, value="Hire month #").font = F_SUB
ws.cell(row=27, column=4, value="Monthly cost").font = F_SUB
for i, (n, m, cst) in enumerate(hires):
    r = 28 + i
    label(ws, r, n)
    c1 = ws.cell(row=r, column=3, value=m); c1.font = F_INPUT; c1.fill = FILL_IN; c1.number_format = NUM
    c2 = ws.cell(row=r, column=4, value=cst); c2.font = F_INPUT; c2.fill = FILL_IN; c2.number_format = CUR

ws.cell(row=33, column=2, value="Fixed operating costs (per month, by year)").font = F_BOLD
ws.cell(row=34, column=3, value="Year 1").font = F_SUB
ws.cell(row=34, column=4, value="Year 2").font = F_SUB
ws.cell(row=34, column=5, value="Year 3").font = F_SUB
opex_lines = [("Software & AI tooling", 450, 700, 900),
              ("Insurance (GL + E&O + cyber)", 290, 320, 380),
              ("Marketing, content & events", 400, 800, 1000),
              ("Travel (client & gov matchmaking)", 250, 600, 800),
              ("Accounting, legal & compliance", 300, 450, 600),
              ("Platform infrastructure", 200, 400, 600),
              ("Phone, mail & miscellaneous", 150, 200, 250)]
for i, (n, a, b, ccc) in enumerate(opex_lines):
    r = 35 + i
    label(ws, r, n)
    for j, v in enumerate((a, b, ccc)):
        c = ws.cell(row=r, column=3 + j, value=v); c.font = F_INPUT; c.fill = FILL_IN; c.number_format = CUR

ws.cell(row=44, column=2, value="Cash & scenarios").font = F_BOLD
arow(45, "Starting cash (owner contribution)", 20000, CUR, "SBA Express (veteran fee relief) available as optional buffer — see Launch Guide")
arow(46, "Tax set-aside (% of positive net income)", 0.25, PCT, "Confirm with CPA")
arow(47, "Conservative scenario multiplier", 0.55, PCT)
arow(48, "Stretch scenario multiplier", 1.40, PCT)

# ---------------------------------------------------------------- Volumes
ws = wb.create_sheet("Volumes")
ws.sheet_properties.tabColor = "1E5BFF"
ws.cell(row=1, column=2, value="VOLUMES — unit plan by month (edit blue cells)").font = F_TITLE
month_header(ws)
def yr(vals12): return vals12
W  = [0,1,1,2,2,2,2,2,2,2,1,1] + [2]*12 + [2]*12
A  = [0,0,1,0,1,0,1,0,1,0,1,0] + [1]*12 + [1,1,2,1,1,2,1,1,2,1,1,1]
P  = [0]*7+[1,0,0,1,0] + [0,1,0,0,1,0,0,1,0,0,1,1] + [0,1,0,1,0,1,1,0,1,1,1,1]
OP = [0]*9+[1,1,2] + [2,2,3,3,3,3,4,4,4,4,5,5] + [5,5,6,6,6,6,7,7,7,7,8,8]
AG = [0,0,0,1,0,0,0,0,1,0,0,0] + [0,0,1,0,0,1,0,0,1,0,0,1] + [0,0,1,0,0,1,0,0,1,0,0,1]
CP = [0,0,0,0,1,1,1,1,1,2,2,2] + [2,3,3,3,4,4,4,4,4,5,5,5] + [5,6,6,6,7,7,7,7,8,8,8,8]
T  = [0,0,0,0,0,1,0,0,0,0,0,0] + [0,0,0,1,0,0,0,1,0,0,0,1] + [0,1,0,1,0,1,0,1,0,1,0,0]
FS = [0]*9+[8000]*3 + [10000]*12 + [14000]*12
FP = [0]*12 + [0]*8+[12000]*4 + [25000]*12
vol_rows = [
 (4,  "Workshops delivered (count)", W, NUM),
 (5,  "Foundation audits started (count)", A, NUM),
 (6,  "Momentum pilots started (count)", P, NUM),
 (7,  "Operating Partner retainers active (count)", OP, NUM),
 (8,  "Agency growth-system builds (count)", AG, NUM),
 (9,  "Agency care plans active (count)", CP, NUM),
 (10, "Training programs delivered (count)", T, NUM),
 (11, "Federal subcontract revenue ($/mo)", FS, CUR),
 (12, "Federal prime revenue ($/mo)", FP, CUR),
]
for r, name, vals, fmt in vol_rows:
    label(ws, r, name)
    put_row(ws, r, vals, F_INPUT, fmt, FILL_IN)
n = ws.cell(row=14, column=2, value="Counts are engagements; federal lines are direct $/month. Pre-filled with the base case — edit freely.")
n.font = F_SUB

# ---------------------------------------------------------------- Revenue
ws = wb.create_sheet("Revenue")
ws.cell(row=1, column=2, value="REVENUE BUILD — calculated from Volumes × Assumptions").font = F_TITLE
month_header(ws); year_headers(ws)
rev_rows = [
 (4,  "Workshops",            lambda i: f"=Volumes!{col(i)}4*Assumptions!$C$5"),
 (5,  "Foundation audits",    lambda i: f"=Volumes!{col(i)}5*Assumptions!$C$6"),
 (6,  "Momentum pilots",      lambda i: (f"=Assumptions!$C$7*(Assumptions!$C$8*Volumes!{col(i)}6)" if i == 1 else
                                          f"=Assumptions!$C$7*(Assumptions!$C$8*Volumes!{col(i)}6+Assumptions!$C$9*Volumes!{col(i-1)}6)")),
 (7,  "Operating Partner retainers", lambda i: f"=Volumes!{col(i)}7*Assumptions!$C$10"),
 (8,  "Agency builds",        lambda i: f"=Volumes!{col(i)}8*Assumptions!$C$11"),
 (9,  "Agency care plans",    lambda i: f"=Volumes!{col(i)}9*Assumptions!$C$12"),
 (10, "Training programs",    lambda i: f"=Volumes!{col(i)}10*Assumptions!$C$13"),
 (11, "Federal subcontracts", lambda i: f"=Volumes!{col(i)}11"),
 (12, "Federal prime work",   lambda i: f"=Volumes!{col(i)}12"),
]
for r, name, f in rev_rows:
    label(ws, r, name)
    put_row(ws, r, f, F_LINK, CUR)
    year_sums(ws, r)
label(ws, 14, "TOTAL REVENUE", bold=True)
put_row(ws, 14, lambda i: f"=SUM({col(i)}4:{col(i)}12)", F_BOLD, CUR)
year_sums(ws, 14, font=F_BOLD)
for i in range(1, MONTHS + 1):
    ws.cell(row=14, column=C0 + i - 1).border = Border(top=Side(style="thin", color="000000"))

# ---------------------------------------------------------------- Opex
ws = wb.create_sheet("Opex")
ws.cell(row=1, column=2, value="OPERATING EXPENSES — calculated from Assumptions").font = F_TITLE
month_header(ws); year_headers(ws)
for k, (nme, _, _) in enumerate(hires):
    r = 4 + k
    label(ws, r, f"Payroll — {nme.lower()}")
    put_row(ws, r, lambda i, rr=28 + k: f"=IF({col(i)}$2>=Assumptions!$C${rr},Assumptions!$D${rr},0)", F_LINK, CUR)
    year_sums(ws, r)
for k in range(len(opex_lines)):
    r = 9 + k
    label(ws, r, opex_lines[k][0])
    put_row(ws, r, lambda i, rr=35 + k:
            f"=CHOOSE(INT(({col(i)}$2-1)/12)+1,Assumptions!$C${rr},Assumptions!$D${rr},Assumptions!$E${rr})",
            F_LINK, CUR)
    year_sums(ws, r)
label(ws, 17, "TOTAL OPERATING EXPENSES", bold=True)
put_row(ws, 17, lambda i: f"=SUM({col(i)}4:{col(i)}15)", F_BOLD, CUR)
year_sums(ws, 17, font=F_BOLD)

# ---------------------------------------------------------------- P&L
ws = wb.create_sheet("P&L")
ws.sheet_properties.tabColor = INK
ws.cell(row=1, column=2, value="PROFIT & LOSS — monthly").font = F_TITLE
month_header(ws); year_headers(ws)
label(ws, 4, "Revenue", bold=True)
put_row(ws, 4, lambda i: f"=Revenue!{col(i)}14", F_LINK, CUR); year_sums(ws, 4)
label(ws, 5, "Delivery costs (COGS)")
put_row(ws, 5, lambda i: f"={col(i)}4*CHOOSE(INT(({col(i)}$2-1)/12)+1,Assumptions!$C$16,Assumptions!$C$17,Assumptions!$C$18)", F_CALC, CUR)
year_sums(ws, 5)
label(ws, 6, "GROSS PROFIT", bold=True)
put_row(ws, 6, lambda i: f"={col(i)}4-{col(i)}5", F_BOLD, CUR); year_sums(ws, 6, font=F_BOLD)
label(ws, 8, "Operating expenses")
put_row(ws, 8, lambda i: f"=Opex!{col(i)}17", F_LINK, CUR); year_sums(ws, 8)
label(ws, 9, "Owner draw")
put_row(ws, 9, lambda i: f"=IF({col(i)}$2>=Assumptions!$C$22,CHOOSE(INT(({col(i)}$2-1)/12)+1,Assumptions!$C$21,Assumptions!$C$23,Assumptions!$C$24),0)", F_CALC, CUR)
year_sums(ws, 9)
label(ws, 11, "NET INCOME (pre-tax)", bold=True)
put_row(ws, 11, lambda i: f"={col(i)}6-{col(i)}8-{col(i)}9", F_BOLD, CUR); year_sums(ws, 11, font=F_BOLD)
label(ws, 12, "Cumulative net income")
put_row(ws, 12, lambda i: (f"={col(i)}11" if i == 1 else f"={col(i-1)}12+{col(i)}11"), F_CALC, CUR)
label(ws, 13, "Cumulative positive? (1/0)")
put_row(ws, 13, lambda i: f"=IF({col(i)}12>0,1,0)", F_CALC, NUM)
for i in range(1, MONTHS + 1):
    ws.cell(row=13, column=C0 + i - 1).font = Font(name="Arial", size=8, color="999999")
ws.cell(row=13, column=2).font = F_SUB

# ---------------------------------------------------------------- Cash Flow
ws = wb.create_sheet("Cash Flow")
ws.sheet_properties.tabColor = INK
ws.cell(row=1, column=2, value="CASH FLOW — monthly").font = F_TITLE
month_header(ws); year_headers(ws)
label(ws, 4, "Beginning cash")
put_row(ws, 4, lambda i: ("=Assumptions!$C$45" if i == 1 else f"={col(i-1)}8"), F_CALC, CUR)
label(ws, 5, "Net income (pre-tax)")
put_row(ws, 5, lambda i: f"='P&L'!{col(i)}11", F_LINK, CUR); year_sums(ws, 5)
label(ws, 6, "Startup costs (one-time)")
put_row(ws, 6, lambda i: ("='Startup Costs'!$C$16*-1" if i == 1 else 0), F_LINK, CUR); year_sums(ws, 6)
label(ws, 7, "Tax set-aside")
put_row(ws, 7, lambda i: f"=-MAX(0,'P&L'!{col(i)}11)*Assumptions!$C$46", F_CALC, CUR); year_sums(ws, 7)
label(ws, 8, "ENDING CASH", bold=True)
put_row(ws, 8, lambda i: f"={col(i)}4+{col(i)}5+{col(i)}6+{col(i)}7", F_BOLD, CUR)
label(ws, 9, "Status")
put_row(ws, 9, lambda i: f'=IF({col(i)}8<0,"SHORTFALL","OK")', F_CALC, "General")

# ---------------------------------------------------------------- Startup Costs
ws = wb.create_sheet("Startup Costs")
ws.sheet_properties.tabColor = "1E5BFF"
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 44
ws.column_dimensions["C"].width = 14; ws.column_dimensions["D"].width = 60
ws.cell(row=2, column=2, value="STARTUP COSTS — one-time (edit blue cells)").font = F_TITLE
items = [
 ("State LLC filing + registered agent (year 1)", 400, "Texas waives filing fees for 100% veteran-owned LLCs — check your state"),
 ("Attorney review — operating agreement (SDVOSB control language)", 750, "Critical for SBA VetCert eligibility"),
 ("Insurance down payments (GL + E&O + cyber)", 600, "Annual premiums modeled monthly in Opex"),
 ("Equipment & workstation", 2500, ""),
 ("Software setup & licenses (one-time)", 500, ""),
 ("Brand & website", 0, "Already complete — aas-fable design system"),
 ("Accounting setup (CPA onboarding, chart of accounts)", 300, ""),
 ("Contracts pack (MSA, SOW, NDA, subcontract templates)", 600, "Attorney-reviewed templates"),
 ("Working capital reserve", 5000, "Buffer before first audit cash lands"),
 ("Contingency", 1000, ""),
]
ws.cell(row=4, column=2, value="Item").font = F_HDR; ws.cell(row=4, column=2).fill = FILL_HDR
ws.cell(row=4, column=3, value="Amount").font = F_HDR; ws.cell(row=4, column=3).fill = FILL_HDR
ws.cell(row=4, column=4, value="Note").font = F_HDR; ws.cell(row=4, column=4).fill = FILL_HDR
for i, (n, v, note) in enumerate(items):
    r = 5 + i
    ws.cell(row=r, column=2, value=n).font = F_LABEL
    c = ws.cell(row=r, column=3, value=v); c.font = F_INPUT; c.fill = FILL_IN; c.number_format = CUR
    ws.cell(row=r, column=4, value=note).font = F_SUB
ws.cell(row=16, column=2, value="TOTAL STARTUP COSTS").font = F_BOLD
c = ws.cell(row=16, column=3, value="=SUM(C5:C14)"); c.font = F_BOLD; c.number_format = CUR

# ---------------------------------------------------------------- Break-Even & Scenarios
ws = wb.create_sheet("Break-Even & Scenarios")
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 46
for cl in "CDE": ws.column_dimensions[cl].width = 16
ws.column_dimensions["F"].width = 56
ws.cell(row=2, column=2, value="BREAK-EVEN").font = F_TITLE
def brow(r, name, formula, fmt=CUR, note="", bold=False):
    label(ws, r, name, bold=bold)
    c = ws.cell(row=r, column=3, value=formula); c.font = F_BOLD if bold else F_CALC; c.number_format = fmt
    if note: ws.cell(row=r, column=6, value=note).font = F_SUB
brow(4, "Average fixed costs / month (Year 1)", "=(SUM(Opex!C17:N17)+SUM('P&L'!C9:N9))/12", CUR, "Opex + owner draw")
brow(5, "Gross margin (Year 1)", "=1-Assumptions!$C$16", PCT)
brow(6, "Required revenue / month to break even", "=C4/C5", CUR, "Roughly one workshop + one-third of an audit", bold=True)
brow(7, "First month cumulative net income turns positive", "=IFERROR(MATCH(1,'P&L'!C13:AL13,0),\"beyond month 36\")", NUM, "Model month # (1 = Jul-26)")
ws.cell(row=10, column=2, value="SCENARIOS — annual revenue and directional net income").font = F_TITLE
hdrs = ["", "Year 1", "Year 2", "Year 3"]
for j, t in enumerate(hdrs):
    c = ws.cell(row=12, column=2 + j, value=t); c.font = F_HDR; c.fill = FILL_HDR
scen = [("Conservative (× multiplier)", "Assumptions!$C$47"),
        ("Base plan (Volumes as entered)", None),
        ("Stretch (× multiplier)", "Assumptions!$C$48")]
for i, (nm, mult) in enumerate(scen):
    r = 13 + i
    ws.cell(row=r, column=2, value=nm).font = F_LABEL
    for k in range(3):
        ref = f"Revenue!{get_column_letter(40 + k)}14"
        f = f"={ref}" if mult is None else f"={ref}*{mult}"
        c = ws.cell(row=r, column=3 + k, value=f); c.font = F_LINK if mult is None else F_CALC; c.number_format = CUR
ws.cell(row=17, column=2, value="Directional net income (revenue × gross margin − fixed costs)").font = F_SUB
for i, (nm, mult) in enumerate(scen):
    r = 18 + i
    ws.cell(row=r, column=2, value=nm).font = F_LABEL
    for k in range(3):
        rev = f"Revenue!{get_column_letter(40 + k)}14" + (f"*{mult}" if mult else "")
        gm = f"(1-Assumptions!$C${16 + k})"
        fixed = f"(Opex!{get_column_letter(40 + k)}17+'P&L'!{get_column_letter(40 + k)}9)"
        c = ws.cell(row=r, column=3 + k, value=f"={rev}*{gm}-{fixed}")
        c.font = F_CALC; c.number_format = CUR
ws.cell(row=22, column=2, value="Scenario rows scale revenue only; cost structure held at base plan. Use for direction, not precision.").font = F_SUB

# ---------------------------------------------------------------- Pipeline
ws = wb.create_sheet("Federal Pipeline")
ws.sheet_properties.tabColor = "1E5BFF"
ws.column_dimensions["A"].width = 3
widths = [34, 22, 18, 14, 10, 14, 18, 30, 14]
for j, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(2 + j)].width = w
ws.cell(row=2, column=2, value="FEDERAL PIPELINE TRACKER").font = F_TITLE
ws.cell(row=3, column=2, value="Sample rows — replace with live opportunities from SAM.gov saved searches, SubNet, and OSDBU forecasts.").font = F_SUB
heads = ["Opportunity", "Agency / Buyer", "Type", "Est. value", "Win %", "Weighted $", "Stage", "Next action", "Due date"]
for j, h in enumerate(heads):
    c = ws.cell(row=5, column=2 + j, value=h); c.font = F_HDR; c.fill = FILL_HDR
samples = [
 ("AI workflow audit — VISN facility ops", "VA (VISN)", "Micro-purchase", 9500, 0.30, "Discovery", "Brief OSDBU small-business specialist", "Aug-26"),
 ("Workflow automation task — prime team", "DoD prime (sub)", "Subcontract", 48000, 0.25, "Teaming", "Send capability statement + past performance", "Sep-26"),
 ("AI adoption training, district office", "USDA", "SAT (<$250K)", 60000, 0.15, "Sources sought", "Respond to RFI; request set-aside", "Oct-26"),
 ("Process intelligence pilot", "GSA", "SDVOSB set-aside", 145000, 0.10, "Watching", "Track on SAM.gov; shape via comments", "Q1-27"),
 ("Ops dashboard modernization", "VA OSDBU lead", "Vets First sole source", 230000, 0.10, "Relationship", "Quarterly check-in; demo platform readout", "Q2-27"),
]
for i, (o, a, t, v, w, st, na, dd) in enumerate(samples):
    r = 6 + i
    vals = [o, a, t, v, w, None, st, na, dd]
    for j, vv in enumerate(vals):
        c = ws.cell(row=r, column=2 + j)
        if j == 5:
            c.value = f"=E{r}*F{r}"; c.font = F_CALC; c.number_format = CUR
        else:
            c.value = vv
            c.font = F_INPUT; c.fill = FILL_IN
            if j == 3: c.number_format = CUR
            if j == 4: c.number_format = PCT
ws.cell(row=12, column=2, value="TOTAL WEIGHTED PIPELINE").font = F_BOLD
c = ws.cell(row=12, column=7, value="=SUM(G6:G10)"); c.font = F_BOLD; c.number_format = CUR

# ---------------------------------------------------------------- Dashboard
ws = wb.create_sheet("Dashboard")
ws.sheet_properties.tabColor = INK
ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 40
for cl in "CDE": ws.column_dimensions[cl].width = 16
ws.cell(row=2, column=2, value="DASHBOARD — model outputs").font = F_TITLE
for j, t in enumerate(["Year 1", "Year 2", "Year 3"]):
    c = ws.cell(row=4, column=3 + j, value=t); c.font = F_HDR; c.fill = FILL_HDR
drows = [
 ("Total revenue", [f"=Revenue!{get_column_letter(40+k)}14" for k in range(3)], CUR),
 ("Gross profit", [f"='P&L'!{get_column_letter(40+k)}6" for k in range(3)], CUR),
 ("Net income (pre-tax, after draw)", [f"='P&L'!{get_column_letter(40+k)}11" for k in range(3)], CUR),
 ("Ending cash (Dec)", [f"='Cash Flow'!{col(12)}8", f"='Cash Flow'!{col(24)}8", f"='Cash Flow'!{col(36)}8"], CUR),
 ("Recurring revenue / month (Dec)", [f"=Revenue!{col(12)}7+Revenue!{col(12)}9", f"=Revenue!{col(24)}7+Revenue!{col(24)}9", f"=Revenue!{col(36)}7+Revenue!{col(36)}9"], CUR),
 ("Federal share of revenue", [f"=(Revenue!{get_column_letter(40+k)}11+Revenue!{get_column_letter(40+k)}12)/Revenue!{get_column_letter(40+k)}14" for k in range(3)], PCT),
]
for i, (nm, fs, fmt) in enumerate(drows):
    r = 5 + i
    label(ws, r, nm)
    for k, f in enumerate(fs):
        c = ws.cell(row=r, column=3 + k, value=f); c.font = F_LINK; c.number_format = fmt
label(ws, 12, "Break-even month (cumulative)")
c = ws.cell(row=12, column=3, value="='Break-Even & Scenarios'!C7"); c.font = F_LINK; c.number_format = NUM

bar = BarChart(); bar.title = "Revenue by year"; bar.style = 10; bar.height = 7; bar.width = 16
data = Reference(ws, min_col=3, max_col=5, min_row=4, max_row=5)
bar.add_data(data, titles_from_data=True, from_rows=False)
bar.legend = None
ws.add_chart(bar, "B15")

wsc = wb["Cash Flow"]
line = LineChart(); line.title = "Ending cash by month"; line.style = 12; line.height = 7; line.width = 30
data = Reference(wsc, min_col=C0, max_col=C0 + MONTHS - 1, min_row=8, max_row=8)
cats = Reference(wsc, min_col=C0, max_col=C0 + MONTHS - 1, min_row=3, max_row=3)
line.add_data(data, titles_from_data=False, from_rows=True)
line.set_categories(cats)
line.legend = None
wsc.add_chart(line, "C12")

wb.save(OUT)
print("Saved:", OUT)
